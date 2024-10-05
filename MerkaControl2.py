# Bloque 1: Importación de librerías y configuración general
# Este bloque contiene las librerías necesarias y la configuración básica para el uso del bot de Telegram y el manejo de archivos Excel.
import logging
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, ConversationHandler, CallbackContext, filters
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Configurar logging para depuración
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

# Bloque 2: Definición de constantes y rutas de archivos
# Define las constantes que se utilizarán en el script, como los estados de la conversación y las rutas absolutas de los archivos Excel.
ASK_NAME = 1  # Estado para pedir el nombre
FILE_PATH = os.path.join(os.getcwd(), 'control_horarios.xlsx')  # Ruta del archivo de control de horarios
USERS_FILE_PATH = os.path.join(os.getcwd(), 'usuarios.xlsx')  # Ruta del archivo de usuarios registrados

# Bloque 3: Creación de archivos Excel (si no existen)
# Este bloque se encarga de crear los archivos `usuarios.xlsx` y `control_horarios.xlsx` en caso de que no existan en la carpeta de trabajo.
def crear_archivo_usuarios():
    """Crea el archivo 'usuarios.xlsx' con las columnas requeridas si no existe."""
    if not os.path.exists(USERS_FILE_PATH):
        logging.info(f"Creando el archivo '{USERS_FILE_PATH}'...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        ws.append(["user_id", "nombre"])  # Encabezados del archivo de usuarios
        wb.save(USERS_FILE_PATH)
        logging.info(f"Archivo '{USERS_FILE_PATH}' creado correctamente.")

def crear_archivo_control():
    """Crea el archivo 'control_horarios.xlsx' con las columnas requeridas si no existe."""
    if not os.path.exists(FILE_PATH):
        logging.info(f"Creando el archivo '{FILE_PATH}'...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Control Horarios"
        ws.append(["Usuario", "Fecha", "Hora de Entrada", "Hora de Salida", "Horas Trabajadas"])  # Encabezados
        wb.save(FILE_PATH)
        logging.info(f"Archivo '{FILE_PATH}' creado correctamente.")

# Bloque 4: Funciones para manejo de usuarios
# Aquí definimos funciones relacionadas con la gestión de usuarios en el archivo `usuarios.xlsx`.
def obtener_nombre_completo(user_id):
    """Obtiene el nombre completo del usuario desde el archivo 'usuarios.xlsx'."""
    try:
        if not os.path.exists(USERS_FILE_PATH):
            logging.warning(f"El archivo {USERS_FILE_PATH} no existe.")
            return None

        wb = load_workbook(USERS_FILE_PATH)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(user_id):
                return row[1]  # Devuelve el nombre completo
        return None
    except Exception as e:
        logging.error(f"Error al obtener nombre completo: {e}")
        return None

def guardar_nombre_completo(user_id, nombre_completo):
    """Guarda el nombre completo del usuario en el archivo 'usuarios.xlsx'."""
    try:
        crear_archivo_usuarios()
        wb = load_workbook(USERS_FILE_PATH)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(user_id):
                for cell in ws["A"]:
                    if cell.value == str(user_id):
                        ws.cell(row=cell.row, column=2, value=nombre_completo)
                        break
                wb.save(USERS_FILE_PATH)
                logging.info(f"Nombre actualizado para el ID {user_id}: {nombre_completo}")
                return

        # Si no existe el usuario, agregar una nueva fila
        ws.append([str(user_id), nombre_completo])
        wb.save(USERS_FILE_PATH)
        logging.info(f"Nombre completo guardado para el ID {user_id}: {nombre_completo}")
    except Exception as e:
        logging.error(f"Error al guardar nombre completo: {e}")

# Bloque 5: Funciones para registro de entrada y salida en 'control_horarios.xlsx'
# Este bloque gestiona las entradas y salidas de los usuarios en el archivo `control_horarios.xlsx`.
def registrar_entrada(usuario, hora_entrada, fecha):
    """Registra la hora de entrada del usuario en 'control_horarios.xlsx'."""
    try:
        crear_archivo_control()
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        ws.append([usuario, fecha, hora_entrada, "", ""])  # Agregar nueva fila con hora de entrada
        wb.save(FILE_PATH)
        logging.info(f"Entrada registrada para {usuario} a las {hora_entrada} el día {fecha}.")
    except Exception as e:
        logging.error(f"Error al registrar la entrada en {FILE_PATH}: {e}")

def registrar_salida(usuario, hora_salida, fecha_salida):
    """Actualiza la hora de salida en la última fila de entrada sin salida registrada para el usuario."""
    try:
        crear_archivo_control()
        wb = load_workbook(FILE_PATH)
        ws = wb.active

        # Buscar la última fila del usuario con "Hora de Salida" vacío
        filas_usuario = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == usuario and not row[3].value:  # Columna 4 es 'Hora de Salida'
                filas_usuario.append(row[0].row)

        if filas_usuario:
            # Obtener la última fila (la más reciente) sin salida registrada
            fila_a_actualizar = max(filas_usuario)

            # Obtener la hora de entrada para calcular las horas trabajadas
            hora_entrada = ws.cell(row=fila_a_actualizar, column=3).value
            fecha_entrada = ws.cell(row=fila_a_actualizar, column=2).value

            # Calcular las horas trabajadas
            try:
                formato = "%H:%M %d/%m/%Y"
                entrada_dt = datetime.strptime(f"{hora_entrada} {fecha_entrada}", formato)
                salida_dt = datetime.strptime(f"{hora_salida} {fecha_salida}", formato)
                horas_trabajadas = salida_dt - entrada_dt
            except ValueError:
                logging.error(f"Error al calcular las horas trabajadas: formato de hora o fecha inválido.")
                return

            # Actualizar la fila con la hora de salida y las horas trabajadas
            ws.cell(row=fila_a_actualizar, column=4, value=hora_salida)
            ws.cell(row=fila_a_actualizar, column=5, value=str(horas_trabajadas))
            wb.save(FILE_PATH)
            logging.info(f"Salida registrada para {usuario} a las {hora_salida}. Horas trabajadas: {horas_trabajadas}.")
        else:
            logging.warning(f"No se encontró una entrada para {usuario} sin salida registrada.")
    except Exception as e:
        logging.error(f"Error al registrar la salida en {FILE_PATH}: {e}")

# Bloque 6: Comandos de Telegram y configuración del bot
# Aquí se definen los comandos `/start`, `/entrada`, `/salida` y la conversación para registrar el nombre completo.

async def start(update: Update, context: CallbackContext) -> None:
    """Comando inicial para dar la bienvenida y explicar el uso del bot."""
    await update.message.reply_text(
        "¡Bienvenido al sistema de control de horario!\n"
        "Usa /registrar para registrar tu nombre completo si es tu primera vez.\n"
        "Luego, usa /entrada para registrar la entrada y /salida para registrar la salida.\n"
        "Puedes ver el reporte de horas trabajadas con /reporte."
    )

async def ask_name(update: Update, context: CallbackContext) -> int:
    """Inicia la conversación para solicitar el nombre completo del usuario."""
    user_id = update.message.from_user.id
    nombre_completo = obtener_nombre_completo(user_id)  # Verificar si el usuario ya está registrado

    if nombre_completo:
        # Si el usuario ya está registrado, le informamos y terminamos la conversación
        await update.message.reply_text(f"Ya estás registrado como {nombre_completo}. Puedes registrar tu entrada con /entrada.")
        return ConversationHandler.END

    # Si no está registrado, pedir el nombre completo
    await update.message.reply_text("Parece que es tu primera vez usando el sistema. Por favor, dime tu nombre completo:")
    return ASK_NAME

async def save_name(update: Update, context: CallbackContext) -> int:
    """Guarda el nombre completo del usuario en el archivo de usuarios."""
    user_id = update.message.from_user.id
    nombre_completo = update.message.text
    guardar_nombre_completo(user_id, nombre_completo)  # Guardar el nombre completo en el archivo de usuarios
    await update.message.reply_text(f"¡Gracias, {nombre_completo}! Ahora puedes registrar tu entrada con /entrada.")
    return ConversationHandler.END

async def cancel(update: Update, context: CallbackContext) -> int:
    """Finaliza la conversación en curso."""
    await update.message.reply_text("La operación ha sido cancelada.")
    return ConversationHandler.END

async def entrada(update: Update, context: CallbackContext) -> None:
    """Registra la entrada de un usuario."""
    try:
        user_id = update.message.from_user.id
        nombre_completo = obtener_nombre_completo(user_id)
        if not nombre_completo:
            await update.message.reply_text("No se ha registrado tu nombre. Usa /registrar para ingresar tu nombre primero.")
            return
        hora_entrada = datetime.now().strftime('%H:%M')
        fecha = datetime.now().strftime('%d/%m/%Y')
        registrar_entrada(nombre_completo, hora_entrada, fecha)
        await update.message.reply_text(f"¡Entrada registrada para {nombre_completo} a las {hora_entrada} el día {fecha}!")
    except Exception as e:
        await update.message.reply_text(f"Error al registrar la entrada: {e}")

async def salida(update: Update, context: CallbackContext) -> None:
    """Registra la salida de un usuario, validando que haya una entrada pendiente."""
    try:
        user_id = update.message.from_user.id
        nombre_completo = obtener_nombre_completo(user_id)
        if not nombre_completo:
            await update.message.reply_text("No se ha registrado tu nombre. Usa /registrar para ingresar tu nombre primero.")
            return

        # Buscar si el usuario tiene una entrada sin salida registrada
        crear_archivo_control()
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        filas_usuario = [row[0].row for row in ws.iter_rows(min_row=2, values_only=False) if row[0].value == nombre_completo and not row[3].value]

        if not filas_usuario:
            # No hay ninguna entrada pendiente de salida, notificar al usuario
            await update.message.reply_text(f"No se encontró ninguna entrada pendiente para {nombre_completo}. Usa /entrada primero.")
            return

        # Si hay una entrada pendiente, registrar la salida
        hora_salida = datetime.now().strftime('%H:%M')
        fecha_salida = datetime.now().strftime('%d/%m/%Y')
        registrar_salida(nombre_completo, hora_salida, fecha_salida)
        await update.message.reply_text(f"¡Salida registrada para {nombre_completo} a las {hora_salida} el día {fecha_salida}!")
    except Exception as e:
        await update.message.reply_text(f"Error al registrar la salida: {e}")




# Bloque 7: Configuración y ejecución del bot
# Se configura el bot, se añaden los manejadores de comandos y se lanza la aplicación.
def run_bot():
    TOKEN = 'TOKEN HERE'
    application = Application.builder().token(TOKEN).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("entrada", entrada))
    application.add_handler(CommandHandler("salida", salida))
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("registrar", ask_name)],
        states={ASK_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_name)]},
        fallbacks=[CommandHandler("cancel", cancel)]
    )
    application.add_handler(conv_handler)
    application.run_polling(stop_signals=None)

# Ejecutar el bot
if __name__ == '__main__':
    run_bot()
